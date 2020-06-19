# tablesToExcel.py

# With Notepad++, use F5 then copy this into box
# C:\Users\Casey\Anaconda3\python.exe -i "$(FULL_CURRENT_PATH)"

# ------------------------------------------------------
# IMPORTS
# ------------------------------------------------------

import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import os
import glob
import string
import copy
from datetime import datetime

# ------------------------------------------------------
# DEFINE FUNCTIONS
# ------------------------------------------------------

# Define writeRow(row, sheet)
def writeRaw(rowText, rowNum, sheet):
	
	# Split input into list
	rowText = rowText.split('\t')
	
	# Loop through list to write to file
	column = 1
	for item in rowText:
		sheet.cell(row=rowNum, column=column).value = item
		sheet.cell(row=rowNum, column=column).alignment = Alignment(horizontal='center', vertical='center')
		fontStyle(sheet.cell(row=rowNum, column=column))
		column += 1
			
# Define writeRow(row, sheet)
def writeExhibit(rowText, rowNum, sheet):
	
	# Split input into list
	rowText = rowText.split('\t')
	
	# Handle screened interval cleanup
	depth = numChop(rowText[7])
	top = numChop(rowText[9])
	bottom = numChop(rowText[10])
	max = numChop(rowText[21])
	
	# Get only columns of interest
	rowText = [rowText[0], rowText[2], rowText[3], rowText[1], rowText[13], rowText[15], rowText[6], depth, top+' - '+bottom, max, rowText[22]] 
	
	# Loop through list to write to file
	column = 1
	for item in rowText:
		if item == 'No Detect Data':
			item = 'NA'
		# Convert dates
		if column == 7 and item != '':
			# Convert to datetime object
			item = datetime.strptime(item[:10], '%Y-%m-%d')
			# Convert to formatted string
			# item = datetime.strftime(item, '%m/%d/%y')
			item = str(item.month) + "/" + str(item.day) + "/" + str(item.year)[-2:]
		if column == 11 and item != 'No Data' and item != 'NA':
			# Convert to datetime object
			item = datetime.strptime(item, '%Y-%m-%d')
			# Convert to formatted string
			# item = datetime.strftime(item, '%m/%d/%y')
			item = str(item.month) + "/" + str(item.day) + "/" + str(item.year)[-2:]
		
		# # Handle depth rounding
		# if column == 5 and item != '':
		
		# if column == 
		sheet.cell(row=rowNum, column=column).value = item
		sheet.cell(row=rowNum, column=column).alignment = Alignment(horizontal='center', vertical='center')
		fontStyle(sheet.cell(row=rowNum, column=column))
		column += 1
	
def fontStyle(cell):
	fontObj1 = Font(name='Times New Roman', size=10)
	cell.font = fontObj1
	cell.border = border
	
def numChop(num):
	if num == 'Multiple' or num == 'No Data' or num == 'No Detect Data':
		return num
	if num != '': 
		try:
			num = str(int(num))
		except:
			num = str(float(num))
			while '.' in num and num[-1] == '0' or num[-1] == '.':
				# if num[-1] == '0':
					# num = num[:-1]
				# elif num[-1] == '.':
					# num = num[:-1]
			# while num[-1] == '0' or num[-1] == '.':
				num = num[:-1]
	return num
	
# ------------------------------------------------------
# INPUTS
# ------------------------------------------------------

# Define path
path = os.path.abspath(os.path.dirname(__file__))

# Define border
border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
	
greyFill = PatternFill(start_color='d3d3d3',
                   end_color='d3d3d3',
                   fill_type='solid')	
				   
# canyonsOut = ['Sandia', 'Mortendad', 'Los Alamos']

# Open the workbook to write to
wb = openpyxl.Workbook()

# Exhibit table
# ------------------------------------------------------

# Open data file
fin = open(path+os.sep+'Tables'+os.sep+'Excluded.txt', 'r')

# Create the sheets to work in
wb.create_sheet(index=0, title='Excluded for Mapping')
wb.create_sheet(index=1, title='Excluded Exhibit')
del wb['Sheet']

# Designate sheet to work in
mappingSheet = wb['Excluded for Mapping']
exhibitSheet = wb['Excluded Exhibit']

# Define the header lines
header = 'Location ID	Aquifer	Latitude	Longitude	Ground Elevation	Geologic Unit	Well Installation Date	Total Well Depth [ft]	Well Diameter [in]	Screen Top [ft]	Screen Bottom [ft]	Comments	WELL_COMPLETION_REPORT_URL	Location Type	Monitoring Area	Watershed	Well Status	Inactive Date	TOC Elevation	LWA Notes	Source	Max Cr	Max Date	Last Cr	Last Date	Max Hex	Max Hex Date	Last Hex	Last Hex Date	Length	Active	Exceedance	Substantial Data'
headerExhibit = 'Well ID	Latitude	Longitude	Aquifer	Type	Watershed	Well Install Date	Well Depth [ft]	Screened Interval [ft]	Max Cr [ug/L]	Date of Max'

# Title
# ------------------------------------------------------
# Define exhibit title
exhibitTitle = 'EXCLUDED LOCATIONS IN THE VICINITY OF THE TA-03 CHROMIUM INVESTIGATION'
# Merge cells for title.
exhibitSheet.merge_cells('A1:K1') 
# Write title to file
exhibitSheet['A1'] = exhibitTitle
# Set format to wrapped
exhibitSheet.cell(row = 1, column = 1).alignment = Alignment(wrap_text = True)
# # Set height
# exhibitSheet.row_dimensions[1].height = 29
# Set font
fontStyle(exhibitSheet['A1'])
# Remove border
exhibitSheet.cell(row = 1, column = 1).border = None

# Write headers to excel sheet
writeRaw(header, 1, mappingSheet)
writeRaw(headerExhibit, 2, exhibitSheet)

# Set height for header row
exhibitSheet.row_dimensions[2].height = 46

# Define rowWidths
rowWidths = [22.15, 10.14, 12, 9.75, 9, 10, 9.15, 7, 13.43, 8.43, 8.43]

# Adjust columns widths and wrap header text
colNum = 0
# Loop through headers
for col in headerExhibit.split('\t'):
	# Set column width
	exhibitSheet.column_dimensions[string.ascii_uppercase[colNum]].width = rowWidths[colNum]
	# Set format to wrapped
	exhibitSheet.cell(row = 2, column = colNum+1).alignment = Alignment(wrap_text = True, horizontal='center', vertical='center')
	colNum += 1
	
rowNum = 2
canyonLast = 'Paj'
# Start extraRow counter
extraRow = 1
# Loop through data in infile
for line in fin:
	# # Define canyon
	# canyon = line.split('\t')[15]
	# # Check if it is a new canyon
	# if canyon != canyonLast and canyon in canyonsOut:
		# # Handle canyons
		# if canyon == canyonsOut[0]:
			# canyonOut = canyonsOut[0]+' Canyon'
		# elif canyon == canyonsOut[1]:
			# canyonOut = canyonsOut[1]+' Canyon'
		# elif canyon == canyonsOut[2]:
			# canyonOut = canyonsOut[2]+' and Pajarito Canyons'
			
		# # Merge cells 
		# exhibitSheet.merge_cells('A'+str(rowNum+extraRow)+':J'+str(rowNum+extraRow)) 
		# # Write canyon to merged cells
		# exhibitSheet['A'+str(rowNum+extraRow)] = canyonOut
		# # Set style
		# exhibitSheet.cell(row = rowNum+extraRow, column = 1).alignment = Alignment(horizontal='left', vertical='center')
		# exhibitSheet.cell(row = rowNum+extraRow, column = 1).fill = greyFill
		# fontStyle(exhibitSheet.cell(row = rowNum+extraRow, column = 1))
		# fontStyle(exhibitSheet.cell(row = rowNum+extraRow, column = 8))
		# extraRow += 1
	# # Update canyonLast
	# canyonLast = canyon
	
	# Write the row to the excel sheet
	writeRaw(line, rowNum, mappingSheet)
	writeExhibit(line, rowNum+extraRow, exhibitSheet)
	rowNum += 1

fin.close()
		
# Change orientation to landscape
openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(exhibitSheet, paper_size = 10, orientation='landscape')

wb.save(path+os.sep+'Tables'+os.sep+'Excluded.xlsx')