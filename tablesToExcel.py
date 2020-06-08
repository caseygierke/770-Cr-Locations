# tablesToExcel.py

# ------------------------------------------------------
# IMPORTS
# ------------------------------------------------------

import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import os
import glob
import string
import copy
from datetime import datetime

# ------------------------------------------------------
# DEFINE FUNCTIONS
# ------------------------------------------------------

# Define writeRow(row, sheet)
def writeRaw(rowText, rowNum, sheet):
	
	# Split input into list
	rowText = rowText.split('\t')
	
	# Loop through list to write to file
	column = 1
	for item in rowText:
		sheet.cell(row=rowNum, column=column).value = item
		sheet.cell(row=rowNum, column=column).alignment = Alignment(horizontal='center', vertical='center')
		fontStyle(sheet.cell(row=rowNum, column=column))
		column += 1
			
# Define writeRow(row, sheet)
def writeExhibit(rowText, rowNum, sheet):
	
	# Split input into list
	rowText = rowText.split('\t')
	
	# Handle screened interval cleanup
	depth = numChop(rowText[7])
	top = numChop(rowText[9])
	bottom = numChop(rowText[10])
	max = numChop(rowText[21])
	
	# Get only columns of interest
	rowText = [rowText[0], rowText[2], rowText[3], rowText[6], depth, top+' - '+bottom, max, rowText[22]] 
	
	# Loop through list to write to file
	column = 1
	for item in rowText:
		# Convert dates
		if column == 4 and item != '':
			# Convert to datetime object
			item = datetime.strptime(item[:10], '%Y-%m-%d')
			# Convert to formatted string
			# item = datetime.strftime(item, '%m/%d/%y')
			item = str(item.month) + "/" + str(item.day) + "/" + str(item.year)[-2:]
		if column == 8 and item != 'No Data':
			# Convert to datetime object
			item = datetime.strptime(item, '%Y-%m-%d')
			# Convert to formatted string
			# item = datetime.strftime(item, '%m/%d/%y')
			item = str(item.month) + "/" + str(item.day) + "/" + str(item.year)[-2:]
		
		# # Handle depth rounding
		# if column == 5 and item != '':
		
		# if column == 
		sheet.cell(row=rowNum, column=column).value = item
		sheet.cell(row=rowNum, column=column).alignment = Alignment(horizontal='center', vertical='center')
		fontStyle(sheet.cell(row=rowNum, column=column))
		column += 1
	
def fontStyle(cell):
	fontObj1 = Font(name='Times New Roman', size=10)
	cell.font = fontObj1
	cell.border = border
	
def numChop(num):
	if num == 'Multiple' or num == 'No Data':
		return num
	if num != '': 
		try:
			num = str(int(num))
		except:
			num = str(float(num))
			while '.' in num and num[-1] == '0' or num[-1] == '.':
				# if num[-1] == '0':
					# num = num[:-1]
				# elif num[-1] == '.':
					# num = num[:-1]
			# while num[-1] == '0' or num[-1] == '.':
				num = num[:-1]
	return num
	
# ------------------------------------------------------
# INPUTS
# ------------------------------------------------------

# Define path
path = os.path.abspath(os.path.dirname(__file__))

# Define border
border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
	
greyFill = PatternFill(start_color='d3d3d3',
                   end_color='d3d3d3',
                   fill_type='solid')	
				   
canyonsOut = ['Sandia', 'Upper Mortendad', 'Lower Mortendad', 'Los Alamos']

# Define aquifers
aquifers = ['Alluvial', 'Intermediate', 'Regional']
# Define file types
types = ['', '- Appendix', '- Excluded']
	
# # Short circuit
# aquifers = [aquifers[0]]
# types = [types[0]]

# Loop through aquifers
for aquifer in aquifers:

	# Loop through file types
	for type in types:
		
		# Get exhibit number based on aquifer
		if aquifer == 'Alluvial' and type == '':
			exNum = 'EXHIBIT CR-4. '
		elif aquifer == 'Alluvial' and type == '- Appendix':
			exNum = 'APPENDIX A. IN'
		elif aquifer == 'Intermediate' and type == '':
			exNum = 'EXHIBIT CR-7. '
		elif aquifer == 'Intermediate' and type == '- Appendix':
			exNum = 'APPENDIX B. IN'
		elif aquifer == 'Regional' and type == '':
			exNum = 'EXHIBIT CR-10. '
		elif aquifer == 'Regional' and type == '- Appendix':
			exNum = 'APPENDIX C. IN'
		
		# if type == '':
		
		# Open the workbook to write to
		wb = openpyxl.Workbook()
		
		# Exhibit table
		# ------------------------------------------------------
		
		# Open data file
		fin = open(path+os.sep+'Tables'+os.sep+aquifer+type+'.txt', 'r')
		
		# Create the sheets to work in
		wb.create_sheet(index=0, title=aquifer+' for Mapping')
		wb.create_sheet(index=1, title=aquifer+' Exhibit')
		del wb['Sheet']
		
		# Designate sheet to work in
		mappingSheet = wb[aquifer+' for Mapping']
		exhibitSheet = wb[aquifer+' Exhibit']

		# Define the header lines
		header = 'Location ID	Aquifer	Latitude	Longitude	Ground Elevation	Geologic Unit	Well Installation Date	Total Well Depth [ft]	Well Diameter [in]	Screen Top [ft]	Screen Bottom [ft]	Comments	WELL_COMPLETION_REPORT_URL	Location Type	Monitoring Area	Watershed	Well Status	Inactive Date	TOC Elevation	LWA Notes	Source	Max Cr	Max Date	Last Cr	Last Date	Max Hex	Max Hex Date	Last Hex	Last Hex Date	Length	Active	Exceedance	Substantial Data'
		headerExhibit = 'Well ID	Latitude	Longitude	Well Install Date	Well Depth [ft]	Screened Interval [ft]	Max Cr [ug/L]	Date of Max'
		
		# Title
		# ------------------------------------------------------
		# Define exhibit title
		exhibitTitle = 'EXHIBIT CR-'+exNum+'. ACTIVE '+aquifer.upper()+' MONITORING WELLS RELATED TO TA-03 CHROMIUM INVESTIGATION'
		# Merge cells for title.
		exhibitSheet.merge_cells('A1:H1') 
		# Write title to file
		exhibitSheet['A1'] = exNum+'ACTIVE '+aquifer.upper()+' MONITORING WELLS RELATED TO TA-03 CHROMIUM INVESTIGATION'
		# Set format to wrapped
		exhibitSheet.cell(row = 1, column = 1).alignment = Alignment(wrap_text = True)
		# Set height
		exhibitSheet.row_dimensions[1].height = 29
		# Set font
		fontStyle(exhibitSheet['A1'])
		# Remove border
		exhibitSheet.cell(row = 1, column = 1).border = None
		
		# Write headers to excel sheet
		writeRaw(header, 1, mappingSheet)
		writeRaw(headerExhibit, 2, exhibitSheet)
		
		# Set height for header row
		exhibitSheet.row_dimensions[2].height = 46
		
		# Define rowWidths
		rowWidths = [11.29, 10.14, 12, 9.15, 7, 13.43, 8.43, 8.43]

		# Adjust columns widths and wrap header text
		colNum = 0
		# Loop through headers
		for col in headerExhibit.split('\t'):
			# Set column width
			exhibitSheet.column_dimensions[string.ascii_uppercase[colNum]].width = rowWidths[colNum]
			# Set format to wrapped
			exhibitSheet.cell(row = 2, column = colNum+1).alignment = Alignment(wrap_text = True, horizontal='center', vertical='center')
			colNum += 1
			
		rowNum = 2
		canyonLast = 'Paj'
		# Start extraRow counter
		extraRow = 1
		# Loop through data in infile
		for line in fin:
			# Define canyon
			canyon = line.split('\t')[15]
			# Check if it is a new canyon
			if canyon != canyonLast and canyon in canyonsOut:
				# Handle canyons
				if canyon == canyonsOut[0]:
					canyonOut = canyonsOut[0]+' Canyon'
				elif canyon == canyonsOut[1] or canyon == 'Lower Mortendad':
					canyonOut = canyonsOut[1][6:]+' Canyon'
				elif canyon == canyonsOut[2]:
					canyonOut = canyonsOut[2]+' and Pajarito Canyons'
					
				# Merge cells 
				exhibitSheet.merge_cells('A'+str(rowNum+extraRow)+':H'+str(rowNum+extraRow)) 
				# Write canyon to merged cells
				exhibitSheet['A'+str(rowNum+extraRow)] = canyonOut
				# Set style
				exhibitSheet.cell(row = rowNum+extraRow, column = 1).alignment = Alignment(horizontal='left', vertical='center')
				exhibitSheet.cell(row = rowNum+extraRow, column = 1).fill = greyFill
				fontStyle(exhibitSheet.cell(row = rowNum+extraRow, column = 1))
				fontStyle(exhibitSheet.cell(row = rowNum+extraRow, column = 8))
				extraRow += 1
			# Update canyonLast
			canyonLast = canyon
			
			# Write the row to the excel sheet
			writeRaw(line, rowNum, mappingSheet)
			writeExhibit(line, rowNum+extraRow, exhibitSheet)
			rowNum += 1
		
		fin.close()
		
		# # ------------------------------------------------------
		# # Appendix table data 
		# # ------------------------------------------------------
		
		# # Open data file
		# fin = open(path+os.sep+'Tables'+os.sep+aquifer+type+'.txt', 'r')
		

		# # Create the sheets to work in
		# wb.create_sheet(index=0, title=aquifer+' for Appendix')
		# wb.create_sheet(index=1, title=aquifer+' Appendix')
		
		# # Designate sheet to work in
		# appMappingSheet = wb[aquifer+' for Appendix']
		# appendixSheet = wb[aquifer+' Appendix']

		# # Define the header lines
		# header = 'Location ID	Aquifer	Latitude	Longitude	Ground Elevation	Geologic Unit	Well Installation Date	Total Well Depth [ft]	Well Diameter [in]	Screen Top [ft]	Screen Bottom [ft]	Comments	WELL_COMPLETION_REPORT_URL	Location Type	Monitoring Area	Watershed	Well Status	Inactive Date	TOC Elevation	LWA Notes	Source	Max Cr	Max Date	Last Cr	Last Date	Max Hex	Max Hex Date	Last Hex	Last Hex Date	Length	Active	Exceedance	Substantial Data'
		# headerExhibit = 'Well ID	Latitude	Longitude	Well Install Date	Well Depth [ft]	Screened Interval [ft]	Max Cr [ug/L]	Date of Max'
		
		# # Title
		# # ------------------------------------------------------
		# # Define exhibit title
		# exhibitTitle = 'EXHIBIT CR-'+exNum+'. ACTIVE '+aquifer.upper()+' MONITORING WELLS RELATED TO TA-03 CHROMIUM INVESTIGATION'
		# # Merge cells for title.
		# exhibitSheet.merge_cells('A1:H1') 
		# # Write title to file
		# exhibitSheet['A1'] = 'EXHIBIT CR-'+exNum+'. ACTIVE '+aquifer.upper()+' MONITORING WELLS RELATED TO TA-03 CHROMIUM INVESTIGATION'
		# # Set format to wrapped
		# exhibitSheet.cell(row = 1, column = 1).alignment = Alignment(wrap_text = True)
		# # Set height
		# exhibitSheet.row_dimensions[1].height = 29
		# # Set font
		# fontStyle(exhibitSheet['A1'])
		# # Remove border
		# exhibitSheet.cell(row = 1, column = 1).border = None
		
		# # Write headers to excel sheet
		# writeRaw(header, 1, mappingSheet)
		# writeRaw(headerExhibit, 2, exhibitSheet)
		
		# # Set height for header row
		# exhibitSheet.row_dimensions[2].height = 46
		
		# # Define rowWidths
		# rowWidths = [11.29, 10.14, 12, 9.15, 7, 11.14, 8.43, 8.43]

		# # Adjust columns widths and wrap header text
		# colNum = 0
		# # Loop through headers
		# for col in headerExhibit.split('\t'):
			# # Set column width
			# exhibitSheet.column_dimensions[string.ascii_uppercase[colNum]].width = rowWidths[colNum]
			# # Set format to wrapped
			# exhibitSheet.cell(row = 2, column = colNum+1).alignment = Alignment(wrap_text = True, horizontal='center', vertical='center')
			# colNum += 1
			
		# rowNum = 2
		# canyonLast = 'Paj'
		# # Start extraRow counter
		# extraRow = 1
		# # Loop through data in infile
		# for line in fin:
			# # Define canyon
			# canyon = line.split('\t')[15]
			# # Check if it is a new canyon
			# if canyon != canyonLast and canyon in canyonsOut:
				# # Handle canyons
				# if canyon == canyonsOut[0]:
					# canyonOut = canyonsOut[0]+' Canyon'
				# elif canyon == canyonsOut[1]:
					# canyonOut = canyonsOut[1][6:]+' Canyon'
				# elif canyon == canyonsOut[2]:
					# canyonOut = canyonsOut[2]+' and Pajarito Canyons'
					
				# # Merge cells 
				# exhibitSheet.merge_cells('A'+str(rowNum+extraRow)+':H'+str(rowNum+extraRow)) 
				# # Write canyon to merged cells
				# exhibitSheet['A'+str(rowNum+extraRow)] = canyonOut
				# # Set style
				# exhibitSheet.cell(row = rowNum+extraRow, column = 1).alignment = Alignment(horizontal='left', vertical='center')
				# exhibitSheet.cell(row = rowNum+extraRow, column = 1).fill = greyFill
				# fontStyle(exhibitSheet.cell(row = rowNum+extraRow, column = 1))
				# fontStyle(exhibitSheet.cell(row = rowNum+extraRow, column = 8))
				# extraRow += 1
			# # Update canyonLast
			# canyonLast = canyon
			
			# # Write the row to the excel sheet
			# writeRaw(line, rowNum, mappingSheet)
			# writeExhibit(line, rowNum+extraRow, exhibitSheet)
			# rowNum += 1
			
		wb.save(path+os.sep+'Tables'+os.sep+aquifer+type+'.xlsx')