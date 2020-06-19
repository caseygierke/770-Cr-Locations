# excludedLocationStats.py

# This plots some summary statistics for excluded Cr locations

# Written by Casey Gierke of Lee Wilson & Associates
# on 5/29/2020

# With Notepad++, use F5 then copy this into box
# C:\Users\Casey\Anaconda3\python.exe -i "$(FULL_CURRENT_PATH)"

# ------------------------------------------------------
# IMPORTS
# ------------------------------------------------------

import pandas as pd
import os
import pyodbc
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl
import numpy as np
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# Adjust size of output plots
font = {'family' : 'Times New Roman',
	'weight' : 'light',
	'size'   : 12}

plt.rc('font', **font)

# ------------------------------------------------------
# DEFINE FUNCTIONS
# ------------------------------------------------------

# Define last position finder
def find_last(s,t):
	last_pos = -1
	while True:
		pos = s.find(t, last_pos +1)
		if pos == -1:
			return last_pos
		last_pos = pos

def DB_get(SQL):
	db = pyodbc.connect("Driver={SQL Server Native Client 11.0};"
						  "Server="+server+";"
						  "Database=Chromium;"
						  "Trusted_Connection=yes;")
	db.autocommit = True

	# Prepare a cursor object using cursor() method
	cursor = db.cursor()
	cursor.execute(SQL)
	result = cursor.fetchall()
	db.close()
	return result

def getCredentials():
	credentialsFile = open(path+os.sep+'Python'+os.sep+'DB Setup'+os.sep+'DB Credentials.txt','r')
	server = credentialsFile.readline()
	if server[-1] == '\n':
		server = server[:-1]
	credentialsFile.close()
	return server

# Define path
path = os.path.abspath(os.path.dirname(__file__))
# Shorten path to one folder up
path = path[:find_last(path,os.sep)]
# Shorten path to one folder up
path = path[:find_last(path,os.sep)]

# Get DB credentials
server = getCredentials()

# DB setup
sql_conn = pyodbc.connect('DRIVER={ODBC Driver 13 for SQL Server};'
							"Server="+server+";"
							'DATABASE=Chromium;'
							'Trusted_Connection=yes') 

# Define writeRow(row, sheet)
def writeExcel(rowText, rowNum, sheet):
	
	# Split input into list
	rowText = rowText.split('\t')
	
	# Loop through list to write to file
	column = 1
	for item in rowText:
		sheet.cell(row=rowNum, column=column).value = item
		sheet.cell(row=rowNum, column=column).alignment = Alignment(horizontal='center', vertical='center')
		fontStyle(sheet.cell(row=rowNum, column=column))
		column += 1
	
def fontStyle(cell):
	fontObj1 = Font(name='Times New Roman', size=10)
	cell.font = fontObj1
	cell.border = border
	
def numChop(num):
	if num == 'Multiple' or num == 'No Data' or num == 'No Detect Data':
		return num
	if num != '': 
		try:
			num = str(int(num))
		except:
			num = str(float(num))
			while '.' in num and num[-1] == '0' or num[-1] == '.':
				# if num[-1] == '0':
					# num = num[:-1]
				# elif num[-1] == '.':
					# num = num[:-1]
			# while num[-1] == '0' or num[-1] == '.':
				num = num[:-1]
	return num
	
# ------------------------------------------------------
# INPUTS
# ------------------------------------------------------

colors = ["white", "blue", "green", "pink", "purple"]

# Define border
border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

# ------------------------------------------------------
# COMPUTATIONS
# ------------------------------------------------------

# All Excluded locations
# ------------------------------------------------------

# Query out all excluded locations
sql = """
		SELECT * FROM
			(SELECT * FROM
				chromium_locations 
			WHERE active IS NULL) AS INACTIVE
		WHERE substantial_data IS NULL
		AND exceedance IS NULL
		ORDER BY aquifer, location_id
		"""

# Query out data
df = pd.read_sql(sql, sql_conn)

# Get number of excluded locations
excludedNum = len(df)

# Get number of excluded locations by aquifer
excludedNumByAquifer = df.groupby(['aquifer']).count().location_id

# Get number of excluded locations by aquifer
excludedNumByAquifer = df.groupby(['aquifer']).count().location_id

# Get number of excluded locations by location_type
excludedNumByType = df.groupby(['location_type']).count().location_id

# ------------------------------------------------------
# Write to excel sheet
# ------------------------------------------------------

# Open the workbook to write to
wb = openpyxl.Workbook()

# Create sheet to work in
wb.create_sheet(index=0, title='All Excluded')
# Delete default worksheet
del wb['Sheet']

# Designate sheet to work in
activeSheet = wb['All Excluded']

# Write total
row = 2
writeExcel('Total Number Excluded\t'+str(excludedNum), 2, activeSheet)
# Set width
activeSheet.column_dimensions['A'].width = 19

# Write headers to excel sheet
writeExcel('Aquifer	Count', 4, activeSheet)
# Write data to excel
row = 5
for aquifer, count in excludedNumByAquifer.items():
	writeExcel(aquifer+'\t'+str(count), row, activeSheet)
	row += 1

# Write headers to excel sheet
row += 1
writeExcel('Type	Count', row, activeSheet)
row += 1

# Write data to excel
for type, count in excludedNumByType.items():
	writeExcel(type+'\t'+str(count), row, activeSheet)
	row += 1
	
# ------------------------------------------------------
# Do pie charts
# ------------------------------------------------------

# Excluded by aquifer
# ------------------------------------------------------
plt.pie(excludedNumByAquifer, labels=pd.Series(df['aquifer'].unique()), colors=colors, autopct='%1.0f%%', wedgeprops={"edgecolor":"k",'linewidth': 1})

# Adjust properties
plt.axis('equal')
plt.title('Excluded data by aquifer type\nof '+str(excludedNum)+' total')
# plt.show()   

# Check that destination directories exist and create if not
if not os.path.exists(path+os.sep+'Python'+os.sep+'Locations'+os.sep+'Excluded'+os.sep):
	os.makedirs(path+os.sep+'Python'+os.sep+'Locations'+os.sep+'Excluded'+os.sep)

# Save the figure
plt.savefig(path+os.sep+'Python'+os.sep+'Locations'+os.sep+'Excluded'+os.sep+'Pie- Aquifer.png',dpi=500)

# Close figure
plt.close('all')

# # Excluded by type
# # ------------------------------------------------------
# # ** Not a good pie chart
# # Get labels
# labels = df['location_type'].unique()
# # Remove None from labels
# if None in labels:
	# labels = np.delete(labels, np.where(labels ==None))

# plt.pie(excludedNumByType, labels=pd.Series(labels), colors=colors, autopct='%1.0f%%', wedgeprops={"edgecolor":"k",'linewidth': 1})

# # Adjust properties
# plt.axis('equal')
# plt.title('Excluded data by location type type\nof '+str(excludedNum)+' total')
# # plt.show()   

# # Save the figure
# plt.savefig(path+os.sep+'Python'+os.sep+'Locations'+os.sep+'Excluded'+os.sep+'Pie- Type.png',dpi=500)
# # Close figure
# plt.close('all')

# ------------------------------------------------------
# Simple bar plot
# ------------------------------------------------------

# Make bar plot
my_plot = excludedNumByType.plot(kind='bar', edgecolor='k') 
# my_plot = sns.barplot(x=excludedNumByType.index, y=excludedNumByType.values, palette=['g','b'], data=excludedNumByType, edgecolor='k');
my_plot.set_ylabel("Number of locations")
my_plot.set_xlabel("")
my_plot.set_title('Excluded location counts by location type')
# Adjust bottom
plt.tight_layout()

# Save the figure
plt.savefig(path+os.sep+'Python'+os.sep+'Locations'+os.sep+'Excluded'+os.sep+'Bar- Type.png',dpi=500)

# plt.show()
# Close figure
plt.close('all')

# ------------------------------------------------------
# Query out totals of active, exceedance and substantial
# ------------------------------------------------------

# Get active count
sql = """SELECT COUNT(*) FROM chromium_locations WHERE active = 'Active'"""
active = DB_get(sql)

# Get exceedance count
sql = """SELECT COUNT(*) FROM chromium_locations WHERE exceedance = 'Exceedance'"""
exceedance = DB_get(sql)

# Get substantial count
sql = """SELECT COUNT(*) FROM chromium_locations WHERE substantial_data = 'Substantial'"""
substantial = DB_get(sql)

# Create sheet to work in
wb.create_sheet(index=0, title='Counts')

# Designate sheet to work in
activeSheet = wb['Counts']

# Write table name
row = 2
writeExcel('Table: Category counts ', row, activeSheet)
activeSheet.cell(row = row, column = 1).alignment = Alignment(horizontal='left', vertical='center')
# Remove border
activeSheet.cell(row = row, column = 1).border = None
		
# Write rows
row += 1
writeExcel('Category\tCount', row, activeSheet)
row += 1
writeExcel('Active\t'+str(active[0][0]), row, activeSheet)
row += 1
writeExcel('Exceedance\t'+str(exceedance[0][0]), row, activeSheet)
row += 1
writeExcel('Substantial\t'+str(substantial[0][0]), row, activeSheet)
row += 1
writeExcel('Total\t'+str(active[0][0] + exceedance[0][0] + substantial[0][0]), row, activeSheet)
row += 1
# Set width
activeSheet.column_dimensions['A'].width = 19

# ------------------------------------------------------
# Query out totals from active category by aquifer
# ------------------------------------------------------

# Get alluvial count
sql = """SELECT COUNT(*) FROM chromium_locations WHERE active = 'Active' and aquifer = 'Alluvial'"""
alluvial = DB_get(sql)

# Get intermediate count
sql = """SELECT COUNT(*) FROM chromium_locations WHERE active = 'Active' and aquifer = 'Intermediate'"""
intermediate = DB_get(sql)

# Get regional count
sql = """SELECT COUNT(*) FROM chromium_locations WHERE active = 'Active' and aquifer = 'Regional'"""
regional = DB_get(sql)

# Get NULL count
sql = """
SELECT COUNT(*) FROM chromium_locations 
WHERE active = 'Active' 
AND aquifer NOT IN ('Alluvial', 'Intermediate', 'Regional') 
OR aquifer IS NULL
"""
null = DB_get(sql)

# Write table name
row += 1
writeExcel('Table: Active counts by aquifer', row, activeSheet)
activeSheet.cell(row = row, column = 1).alignment = Alignment(horizontal='left', vertical='center')
# Remove border
activeSheet.cell(row = row, column = 1).border = None

# Write rows
row += 1
writeExcel('Aquifer\tCount', row, activeSheet)
row += 1
writeExcel('Alluvial\t'+str(alluvial[0][0]), row, activeSheet)
row += 1
writeExcel('Intermediate\t'+str(intermediate[0][0]), row, activeSheet)
row += 1
writeExcel('Regional\t'+str(regional[0][0]), row, activeSheet)
row += 1
writeExcel('Not Identified\t'+str(null[0][0]), row, activeSheet)
row += 1
writeExcel('Total\t'+str(alluvial[0][0] + intermediate[0][0] + regional[0][0] + null[0][0]), row, activeSheet)
row += 1

# ------------------------------------------------------
# Query out totals from exceedance/substantial category by aquifer
# ------------------------------------------------------

# Get alluvial count
sql = """
SELECT COUNT(*) FROM
	(SELECT * FROM
		(SELECT * FROM chromium_locations 
		WHERE active IS NULL) AS INACTIVE
	WHERE exceedance = 'Exceedance'
	OR substantial_data = 'Substantial') AS Exceedance
WHERE aquifer = 'Alluvial'"""
alluvial = DB_get(sql)

# Get intermediate count
sql = """
SELECT COUNT(*) FROM
	(SELECT * FROM
		(SELECT * FROM chromium_locations 
		WHERE active IS NULL) AS INACTIVE
	WHERE exceedance = 'Exceedance'
	OR substantial_data = 'Substantial') AS Exceedance
WHERE aquifer = 'Intermediate'"""
intermediate = DB_get(sql)

# Get regional count
sql = """
SELECT COUNT(*) FROM
	(SELECT * FROM
		(SELECT * FROM chromium_locations 
		WHERE active IS NULL) AS INACTIVE
	WHERE exceedance = 'Exceedance'
	OR substantial_data = 'Substantial') AS Exceedance
WHERE aquifer = 'Regional'"""

regional = DB_get(sql)

# Get NULL count
sql = """
SELECT COUNT(*) FROM
	(SELECT * FROM
		(SELECT * FROM chromium_locations 
		WHERE active IS NULL) AS INACTIVE
	WHERE exceedance = 'Exceedance'
	OR substantial_data = 'Substantial') AS Exceedance
WHERE aquifer NOT IN ('Alluvial', 'Intermediate', 'Regional') 
OR aquifer IS NULL"""
null = DB_get(sql)

# Write table name
row += 1
writeExcel('Table: Exceedances and substantial data counts by aquifer', row, activeSheet)
activeSheet.cell(row = row, column = 1).alignment = Alignment(horizontal='left', vertical='center')
# Remove border
activeSheet.cell(row = row, column = 1).border = None

# Write rows
row += 1
writeExcel('Aquifer\tCount', row, activeSheet)
row += 1
writeExcel('Alluvial\t'+str(alluvial[0][0]), row, activeSheet)
row += 1
writeExcel('Intermediate\t'+str(intermediate[0][0]), row, activeSheet)
row += 1
writeExcel('Regional\t'+str(regional[0][0]), row, activeSheet)
row += 1
writeExcel('Not Identified\t'+str(null[0][0]), row, activeSheet)
row += 1
writeExcel('Total\t'+str(alluvial[0][0] + intermediate[0][0] + regional[0][0] + null[0][0]), row, activeSheet)
row += 1

# ------------------------------------------------------
# Query out totals from exceedance/substantial category by aquifer
# ------------------------------------------------------

sql = """
SELECT * FROM
	(SELECT * FROM
	chromium_locations 
	WHERE active IS NULL) AS INACTIVE
WHERE substantial_data IS NULL
AND exceedance IS NULL
AND max NOT IN ('No Data', 'No Detect Data')
ORDER BY max DESC, location_id
"""

# ------------------------------------------------------
# Simple bar plot of excluded data with data by type
# ------------------------------------------------------

# Query out data
df = pd.read_sql(sql, sql_conn)

* None type does not get included in grouping
*Change it to NULL before grouping

# Get number of excluded locations by location_type
excludedNumByType = df.groupby(['location_type']).count().location_id

# Make bar plot
my_plot = excludedNumByType.plot(kind='bar', edgecolor='k') 
# my_plot = sns.barplot(x=excludedNumByType.index, y=excludedNumByType.values, palette=['g','b'], data=excludedNumByType, edgecolor='k');
my_plot.set_ylabel("Number of locations")
my_plot.set_xlabel("")
my_plot.set_title('Excluded location counts by location type')
# Adjust bottom
plt.tight_layout()

# Save the figure
plt.savefig(path+os.sep+'Python'+os.sep+'Locations'+os.sep+'Excluded'+os.sep+'Bar- Excluded Type Data.png',dpi=500)

# Write table name
row += 1
writeExcel('Table: Excluded wells with data by type', row, activeSheet)
activeSheet.cell(row = row, column = 1).alignment = Alignment(horizontal='left', vertical='center')
# Remove border
activeSheet.cell(row = row, column = 1).border = None

# Write headers to excel sheet
row += 1
writeExcel('Type	Count', row, activeSheet)
row += 1

# Write data to excel
for type, count in excludedNumByType.items():
	writeExcel(type+'\t'+str(count), row, activeSheet)
	row += 1

# """
# SELECT * FROM
	# (SELECT * FROM chromium_locations 
	# WHERE active IS NULL) AS INACTIVE
# WHERE aquifer NOT IN ('Alluvial', 'Intermediate', 'Regional') 
# OR aquifer IS NULL
# """

# Save workbook
wb.save(path+os.sep+'Python'+os.sep+'Locations'+os.sep+'Excluded'+os.sep+'Excluded Stats.xlsx')
